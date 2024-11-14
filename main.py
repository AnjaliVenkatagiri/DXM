import queue
import re
from datetime import timedelta, datetime
import requests
from sheetfu import SpreadsheetApp, Table
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver import ChromeOptions as Options, Keys
from selenium.webdriver import ChromeService as Service
from subprocess import CREATE_NO_WINDOW
from tkinter import messagebox
import tkinter as tk
import traceback
import json
from threading import Thread
# from openpyxl import load_workbook
from queue import Queue
import openpyxl
# from webdriver_manager.chrome import ChromeDriverManager

status_queue = Queue()


def give_error(error_title, error):
    th = Thread(target=lambda: messagebox.showerror(error_title, error), daemon=True)
    th.start()


def data_exists(data):
    sa = SpreadsheetApp('botsservers-3c9f1e68cf1e.json')
    spreadsheet = sa.open_by_id('15K3nyMLbzwGtYBj69M3ElUdjWt8jkSaZVP5DtZdif4w')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    table = Table(data_range)
    for item in table:
        if data["CONo"] == item.get_field_value("CONo"):
            print("Data Exists")
            return True
    return False


def log(data):
    status_queue.put(data)
    # wb = load_workbook("DXMLog.xlsx")
    # ws = wb.active
    # date = datetime.now().strftime("%d-%m-%Y %H:%M.%S")
    # ws.cell(ws.max_row + 1, 1).value = data
    # ws.cell(ws.max_row + 1, 2).value = date
    # wb.save("DXMLog.xlsx")
    # wb.close()


def add_data(data):
    print(data)
    sa = SpreadsheetApp('botsservers-3c9f1e68cf1e.json')
    spreadsheet = sa.open_by_id('15K3nyMLbzwGtYBj69M3ElUdjWt8jkSaZVP5DtZdif4w')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    table = Table(data_range)
    table.add_one(data)
    table.commit()


def update_hsn(item_no):
    service = Service()
    options = Options()
    # options.add_experimental_option('detach', True)
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 10)
    action = webdriver.ActionChains(driver)
    driver.get("http://intranetn.shahi.co.in:8080/ShahiExportIntranet/login")
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.send_keys('755921')
    password = driver.find_element(By.ID, "password")
    password.send_keys('vis1234')
    driver.find_element(By.ID, 'savebutton').click()
    time.sleep(2)
    handles = [driver.current_window_handle]
    driver.execute_script(
        "javascript:openMenuPage('null' , 'SRM    Workplace......GST' , '2452' , 'F' , 'Applications'  );")
    time.sleep(2)
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    handles.append(driver.current_window_handle)
    driver.execute_script(
        f"javascript:openAccessPage('http://f1pla-srm01.shahi.co.in:8080/SRMSCMPROD/mksessSessionAction' , 'SRM Workplace' , '2455' , 'R' , '751203' , 'Y', '50015135', 'N');")
    time.sleep(2)
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    time.sleep(5)
    span_list = driver.find_elements(By.TAG_NAME, "span")
    for i in span_list:
        if "340" in i.get_attribute("innerText"):
            i.click()
            break
    wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, "a[title = 'PPS220']"))).click()
    frame = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, "iframe[src *= 'HSNUpdateAction']")))
    driver.switch_to.frame(frame)
    item_input = driver.find_element(By.ID, "SEARCH_STYLE")
    driver.execute_script(f"arguments[0].value = '{item_no}';", item_input)
    searchbtn = driver.find_element(By.ID, "searchbtn")
    driver.execute_script("arguments[0].click();", searchbtn)
    time.sleep(3)
    hsn_select = driver.find_element(By.XPATH,
                                     '//*[@id="hsnupdate_wrapper"]/div/div[1]/div/table/thead/tr/th[13]/div/a')
    driver.execute_script("arguments[0].click();", hsn_select)
    time.sleep(1)
    new_frame = driver.find_element(By.ID, "handlesuppiframe")
    driver.switch_to.frame(new_frame)
    hsn_ip = driver.find_element(By.ID, "HSN_CODE")
    driver.execute_script("arguments[0].value = '62034290-5'", hsn_ip)
    searchbtn = driver.find_element(By.ID, "searchbtn")
    driver.execute_script("arguments[0].click();", searchbtn)
    time.sleep(1)
    img = driver.find_element(By.XPATH, '//*[@id="dataTable"]/tbody/tr[2]/td[4]/img')
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    check = driver.find_element(By.ID, "CHK_ALL")
    driver.execute_script("arguments[0].click();", check)
    update_btn = driver.find_element(By.ID, "updatebtn")
    driver.execute_script("arguments[0].click();", update_btn)
    driver.switch_to.alert.accept()
    time.sleep(20)


def update_CO(co, po, sp):
    if not co.startswith('20'):
        print(f"Skipping CO {co} as it doesn't start with '20'")
        return
    service = Service()
    options = Options()
    # options.add_experimental_option('detach', True)
    driver = webdriver.Chrome(service=service, options=options)
    excel_file_path = 'credentials1.xlsx'
    wait = WebDriverWait(driver, 30)
    action = webdriver.ActionChains(driver)
    driver.get("http://intranetn.shahi.co.in:8080/ShahiExportIntranet/login")
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    user_id = sheet['A2'].value  # Assuming username is in cell A1
    password_value = sheet['B2'].value
    username.send_keys(user_id)
    password = driver.find_element(By.ID, "password")
    password.send_keys(password_value)
    driver.find_element(By.ID, 'savebutton').click()
    time.sleep(2)
    handles = [driver.current_window_handle]
    driver.execute_script("javascript:openMenuPage('null' , 'CRM - Work Place (New M3)' , '2447' , 'F' , 'Applications'  );")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    handles.append(driver.current_window_handle)
    time.sleep(2)
    driver.execute_script(
        # f"javascript:openAccessPage('http://intranet.shahi.co.in:8080/IntraNet/CRMPRDNEW.jsp' , 'CRM' , '2448' , 'R' , '{user_id}' , 'N', '50012556', 'null');")
        f"javascript:openAccessPage('http://crmm4.shahi.co.in:8080/CRMPRDN/CRMPRDNEW.jsp' , 'CRM' , '2448' , 'R' , '{user_id}' , 'N', '50004167', 'N');")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    time.sleep(10)
    driver.execute_script("getHome('340')")
    frame = driver.find_element(By.ID, "mainFrame")
    driver.switch_to.frame(frame)
    apps = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="mainContainer"]/div[1]')))
    all_apps = apps.find_elements(By.TAG_NAME, "span")
    for i in all_apps:
        if "Sampling Orders" in i.get_attribute("innerText"):
            # i.click()
            driver.execute_script("arguments[0].click();", i)
            break
    frame = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="Samplingntainer"]/iframe')))
    driver.switch_to.frame(frame)
    co_search = wait.until(ec.presence_of_element_located((By.ID, "CO_SEARCH")))
    driver.execute_script(f"arguments[0].value = '{co}';", co_search)
    list_btn = driver.find_element(By.ID, "listorderbtn")
    driver.execute_script("arguments[0].click();", list_btn)
    time.sleep(5)
    edit_btn = wait.until(ec.presence_of_element_located((By.XPATH, "/html/body/form/table/tbody/tr[2]/td/div/table/tbody/tr/td[15]/a[1]/img")))
    driver.execute_script("arguments[0].click();", edit_btn)
    bpo_entry = wait.until(ec.presence_of_element_located((By.ID, "BUYER_PO")))
    driver.execute_script(f"arguments[0].value = '{po}';", bpo_entry)
    pac_term = driver.find_element(By.ID, "PACK_TERMAUTO")
    driver.execute_script("arguments[0].value = 'LSE-LOOSE PACK';", pac_term)
    del_method = driver.find_element(By.ID, "DEL_METHOD")
    driver.execute_script("arguments[0].value = 'AIR';", del_method)
    del_term = driver.find_element(By.ID, "DEL_TERMS")
    driver.execute_script("arguments[0].value = 'C&F';", del_term)
    pay_method = driver.find_element(By.ID, "PAY_METHODAUTO")
    driver.execute_script("arguments[0].value = 'BTR-Bank Transfer';", pay_method)
    pay_term = driver.find_element(By.ID, "PAY_TERMSAUTO")
    driver.execute_script("arguments[0].value = '082-TT  60 Days';", pay_term)
    sale_price = driver.find_element(By.ID, "sapr")
    driver.execute_script(f"arguments[0].value = '{sp}';", sale_price)
    time.sleep(2)
    del_address = driver.find_element(By.XPATH, '//*[@id="BUY_ADDRESS"]')
    driver.execute_script("arguments[0].value = '10';", del_address)
    try:
        save_btn = driver.find_element(By.ID, "savebtn")
        driver.execute_script("arguments[0].click();", save_btn)
    except:
        log("Already Created CO")


def plan_creation(co):
    service = Service()
    options = Options()
    # options.add_experimental_option('detach', True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Chrome(service=service, options=options)
    excel_file_path = 'credentials1.xlsx'
    wait = WebDriverWait(driver, 10)
    action = webdriver.ActionChains(driver)
    driver.get("http://intranetn.shahi.co.in:8080/ShahiExportIntranet/login")
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    user_id = sheet['A2'].value  # Assuming username is in cell A1
    password_value = sheet['B2'].value
    username.send_keys(user_id)
    password = driver.find_element(By.ID, "password")
    password.send_keys(password_value)
    driver.find_element(By.ID, 'savebutton').click()
    time.sleep(2)
    handles = [driver.current_window_handle]
    driver.execute_script("javascript:openMenuPage('null' , 'LG - Shipment Module - New M3......GST' , '2806' , 'F' , 'Applications'  );")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    handles.append(driver.current_window_handle)
    driver.execute_script(f"javascript:openAccessPage('http://f1pla-production02.shahi.co.in:8080/ShahiLogisticsDiv/shahiwebpages/MvxExp/PRE/division.action?appName=LGSHP' , 'Logistics - Shipment Module - Division' , '12830' , 'R' , '{user_id}' , 'N', '50012556', 'null');")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    time.sleep(5)
    spans = driver.find_elements(By.TAG_NAME, "span")
    for i in spans:
        if "340" in i.get_attribute("innerText"):
            driver.execute_script("arguments[0].click();", i)
            break
    links = driver.find_elements(By.TAG_NAME, "a")
    for i in links:
        if "Shipment Plan - Garment" in i.get_attribute("innerText"):
            driver.execute_script("arguments[0].click();", i)
            break
    frame = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'iframe[src *= "ShipmentPlanningNew"]')))
    driver.switch_to.frame(frame)
    new = driver.find_element(By.ID, "NEWsearchId")
    driver.execute_script("arguments[0].click();", new)
    time.sleep(2)
    ship_state = driver.find_element(By.XPATH, '//*[@id="LOCATION_CODE"]')
    ship_state.click()
    ship_state.send_keys("340")
    ship_state.send_keys(Keys.ENTER)
    time.sleep(5)
    unit = driver.find_element(By.XPATH, '//*[@id="planUnit"]')
    unit.click()
    driver.execute_script("arguments[0].value = 'A7'", unit)
    unit.send_keys(Keys.ENTER)
    co_input = wait.until(ec.presence_of_element_located((By.ID, "S_CO_NUMB")))
    driver.execute_script(f"arguments[0].value = '{co}'", co_input)
    search = driver.find_element(By.XPATH,
                                 '/html/body/form/div[1]/div[2]/div[1]/div[2]/table[2]/tbody/tr[1]/td[4]/div/img')
    driver.execute_script("arguments[0].click();", search)
    try:
        save = wait.until(ec.presence_of_element_located((By.ID, "saveId")))
        unit = driver.find_element(By.ID, "planUnit")
        driver.execute_script("arguments[0].value = 'A7';", unit)
        driver.execute_script("arguments[0].click();", save)
        time.sleep(5)
        driver.switch_to.alert.accept()
        time.sleep(10)
        driver.switch_to.alert.accept()
        plan = driver.find_element(By.ID, "PLAN_NUMB")
    except:
        plan = driver.find_element(By.XPATH, "/html/body/form/div[1]/div[3]/div[1]/ul/li/span")
        plan_text = plan.get_attribute("innerText")
        return re.findall(r"\b\d{7}\b", plan_text)[0]
    return plan.get_attribute("value")


def invoice(plan, upcharval, sales_price, description):
    date_format = datetime.now().strftime("%d-%m-%Y")
    prefs = {"download.default_directory": fr"C:\InvoicesDownload-{date_format}"}
    service = Service()
    options = Options()
    options.add_experimental_option("prefs", prefs)
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Chrome(options=options, service=service)
    wait = WebDriverWait(driver, 10)
    action = webdriver.ActionChains(driver)
    driver.get("http://intranetn.shahi.co.in:8080/ShahiExportIntranet/login")
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.send_keys('734129')
    password = driver.find_element(By.ID, "password")
    password.send_keys('Suni@734129')
    submit = driver.find_element(By.ID, "savebutton")
    submit.click()
    handles = []
    handle = driver.current_window_handle
    handles.append(handle)
    dropdown = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="bs-example-navbar-collapse-1"]/div/div/ul/li[2]/a/span[1]')))
    # dropdown.click()
    driver.execute_script(
        "javascript:openMenuPage('null' , 'LG - Shahi Logistics Pre Shipment' , '10360' , 'F' , 'Applications'  );")
    all_handles = driver.window_handles
    for i in all_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    handles.append(driver.current_window_handle)
    driver.execute_script(
        f"javascript:openAccessPage('http://f1pla-production02.shahi.co.in:8080/ShahiLogisticsDiv/shahiwebpages/MvxExp/PRE/division.action?appName=LGPRE' , 'Logistic Logistics Pre Shipment    -  (Division)' , '12828' , 'R' , '734129' , 'N', '50008181', 'N');")
    for i in driver.window_handles:
        if i not in handles:
            driver.switch_to.window(i)
            break
    time.sleep(5)
    table = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="main-content"]/div/div/div[2]/table')))
    spans = table.find_elements(By.TAG_NAME, "span")
    for i in spans:
        if "340" in i.get_attribute("innerText"):
            i.click()
            break
    ul = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, "#tabs > ul")))
    a_list = ul.find_elements(By.TAG_NAME, "a")
    for i in a_list:
        if "Invoice Pre Shipment" in i.get_attribute("innerText"):
            i.click()
            break
    frame = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'iframe[src *= "PREINVMVX"]')))
    driver.switch_to.frame(frame)
    plan_input = driver.find_element(By.ID, "searchplan")
    driver.execute_script(f"arguments[0].value = '{plan}';", plan_input)
    go = driver.find_element(By.ID, "searchheadId")
    driver.execute_script("arguments[0].click();", go)
    invoice_type = wait.until(ec.presence_of_element_located((By.ID, "self_tp")))
    driver.execute_script("arguments[0].value = 'S';", invoice_type)
    pch = driver.find_element(By.ID, "cost_centre")
    driver.execute_script("arguments[0].value = 'MNB';", pch)
    exp_type = driver.find_element(By.ID, "exp_type")
    driver.execute_script("arguments[0].value = 'GMN';", exp_type)
    lut = driver.find_element(By.ID, "LUT_IGST")
    driver.execute_script("arguments[0].value = 'LUT';", lut)
    ship_mode = driver.find_element(By.ID, "mode_of_ship")
    driver.execute_script("arguments[0].value = 'COU';", ship_mode)
    pre_carriage = driver.find_element(By.ID, "pre_carriage")
    driver.execute_script('arguments[0].value = "ROAD/B\'LORE";', pre_carriage)
    ac_holder = driver.find_element(By.ID, "ac_holder")
    driver.execute_script("arguments[0].value = 'NATARAJ';", ac_holder)
    merchant = driver.find_element(By.ID, "merchant")
    driver.execute_script("arguments[0].value = 'PRASHANTH HIREMATH';", merchant)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[5]/td[4]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'TPC00053';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[5]/td[4]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'TPC00053';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[6]/td[4]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'TPC00053';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    ship_term = driver.find_element(By.ID, "ship_term")
    driver.execute_script("arguments[0].value = 'C&F';", ship_term)
    pay_term = driver.find_element(By.ID, "payment_term")
    driver.execute_script("arguments[0].value = 'TT1';", pay_term)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[7]/td[4]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'A7';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[2]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr[1]/td[4]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    inv_date = driver.find_element(By.ID, "inv_date")
    inv_date = inv_date.get_attribute("value")
    date = datetime.strptime(inv_date, "%d-%b-%Y")
    inv_date = date.strftime("%Y-%m-%d")
    date = date + timedelta(days=7)
    date_str = date.strftime("%d/%m/%Y")
    etd_date = driver.find_element(By.ID, "etd_date")
    driver.execute_script(f"arguments[0].value = '{date_str}';", etd_date)
    desc = driver.find_element(By.ID, "SHIP_DESC")
    driver.execute_script("arguments[0].value = 'UPCHARGE UPDATED TOWARDS COURIER FREIGHT CHARGE';", desc)
    upcharge = driver.find_element(By.ID, "upcharge_per")
    driver.execute_script(f"arguments[0].value = '{upcharval}'", upcharge)
    remarks = driver.find_element(By.ID, 'remarks')
    driver.execute_script("arguments[0].value = 'Generating invoice'", remarks)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'BANGALORE';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[2]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'BLR4';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[3]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'BLR4';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[4]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'NYK';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[5]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'US';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    img = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[6]/td[10]/a/img')
    driver.execute_script("arguments[0].click();", img)
    handle_frm = driver.find_element(By.ID, "handlefrm")
    driver.switch_to.frame(handle_frm)
    unitprm = driver.find_element(By.ID, "unitparam")
    driver.execute_script("arguments[0].value = 'IN';", unitprm)
    img = driver.find_element(By.XPATH, '/html/body/form/div/div[1]/div/table/tbody/tr/td[3]/img')
    driver.execute_script("arguments[0].click();", img)
    img = wait.until(ec.presence_of_element_located(
        (By.XPATH, '/html/body/form/div/div[2]/div/table/tbody/tr/td/div/table/tbody/tr[1]/td[3]/img')))
    driver.execute_script("arguments[0].click();", img)
    driver.switch_to.default_content()
    driver.switch_to.frame(frame)
    dest = driver.find_element(By.ID, "DESTI_CODE_DESC")
    driver.execute_script("arguments[0].value = 'USA';", dest)
    dest_cntry = driver.find_element(By.ID, "DESTI_CNTRY_DESC")
    driver.execute_script("arguments[0].value = 'UNITED STATES OF AMERICA';", dest_cntry)
    fob = driver.find_element(By.XPATH, '//*[@id="tab1"]/table/tbody/tr[1]/td/table/tbody/tr[9]/td[10]/input[1]')
    driver.execute_script(f"arguments[0].value = '{sales_price}';", fob)
    net_fob = driver.find_element(By.CSS_SELECTOR, 'input[name = "NET_FOB"]')
    driver.execute_script(f"arguments[0].value = '{sales_price}';", net_fob)
    tab_bar = driver.find_element(By.CSS_SELECTOR, 'a[title = "Billing Details"]')
    driver.execute_script("arguments[0].click();", tab_bar)
    category = driver.find_element(By.ID, "CATG_CODE_COPY")
    driver.execute_script("arguments[0].value = '6';", category)
    desc = driver.find_element(By.ID, "CATG_DESC_COPY")
    driver.execute_script(f'arguments[0].value = "{description}";', desc)
    ship_type = driver.find_element(By.ID, "SHIP_TYPE_COPY")
    driver.execute_script("arguments[0].value = 'SAMPLE';", ship_type)
    scheme = driver.find_element(By.ID, "SCHEME_CODE_COPY")
    driver.execute_script("arguments[0].value = '00';", scheme)
    # img = driver.find_element(By.XPATH, '//*[@id="dbkslhref"]/img')
    # driver.execute_script("arguments[0].click();", img)
    # handle_frm = driver.find_element(By.ID, "handlefrm")
    # driver.switch_to.frame(handle_frm)
    # img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[2]/img')
    # driver.execute_script("arguments[0].click();", img)
    # img = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[4]/img')))
    # driver.execute_script("arguments[0].click();", img)
    # driver.switch_to.default_content()
    # driver.switch_to.frame(frame)
    # img = driver.find_element(By.XPATH, '//*[@id="roslslhref"]/img')
    # driver.execute_script("arguments[0].click();", img)
    # handle_frm = driver.find_element(By.ID, "handlefrm")
    # driver.switch_to.frame(handle_frm)
    # unitprm = driver.find_element(By.ID, "unitparam")
    # img = driver.find_element(By.XPATH, '/html/body/form/div/div[2]/div[1]/table/tbody/tr/td[2]/img')
    # driver.execute_script("arguments[0].click();", img)
    # img = wait.until(ec.presence_of_element_located(
    #     (By.XPATH, '/html/body/form/div/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[4]/img')))
    # driver.execute_script("arguments[0].click();", img)
    # driver.switch_to.default_content()
    # driver.switch_to.frame(frame)
    copy = driver.find_element(By.CSS_SELECTOR, "input[value =  'CP All']")
    driver.execute_script("arguments[0].click();", copy)
    save = driver.find_element(By.ID, "saveheadId")
    driver.execute_script("arguments[0].click();", save)
    try:
        driver.switch_to.alert.accept()
        ci = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div[1]/div[2]/div[2]/div/div/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/ul/li/span')))
        ci_text = ci.get_attribute("innerText")
        ci_no = re.findall(r"\d{10}", ci_text)[0]
        search_CI = driver.find_element(By.ID, "searchcino")
        driver.execute_script(f"arguments[0].value = '{ci_no}';", search_CI)
    except:
        ci_no = driver.find_element(By.ID, "searchcino").get_attribute("value")
    go = driver.find_element(By.ID, "searchheadId")
    driver.execute_script("arguments[0].click();", go)
    try:
        forward = wait.until(ec.presence_of_element_located((By.ID, "fwd_custom")))
        driver.execute_script("arguments[0].click();", forward)
        time.sleep(1)
        save = driver.find_element(By.ID, "saveId")
        driver.execute_script("arguments[0].click();", save)
        driver.switch_to.alert.accept()
        inv = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div[1]/div[2]/div[2]/div/div/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/ul/li/span'))).get_attribute("innerText")
        inv_no = re.findall(r"\d{12}", inv)[0]
        inv_search = driver.find_element(By.ID, "searchinv")
        driver.execute_script(f"arguments[0].value = '{inv_no}';", inv_search)
        search_btn = driver.find_element(By.ID, "searchId")
        driver.execute_script("arguments[0].click();", search_btn)
    except:
        search_btn = driver.find_element(By.ID, "searchId")
        driver.execute_script("arguments[0].click();", search_btn)
        inv_search = driver.find_element(By.ID, "searchinv")
        inv_no = inv_search.get_attribute("value")
    print_inv = wait.until(ec.presence_of_element_located((By.ID, "printheadId")))
    driver.execute_script("arguments[0].click();", print_inv)
    time.sleep(5)
    return inv_date, ci_no, inv_no, None


def bot():
    log("Bot Started")
    date = datetime.now().strftime("%d/%m/%Y")
    # date = "07/09/2023"
    response = requests.get(
        f"https://careers.shahi.co.in/shahiapiprods/apiTest/getOrder?fdate=01/09/2023&tdate={date}")
    print(response)
    response = response.json()
    print(response)
    log("Fetched Response From API")
    # for i, row in enumerate(response):
    #     if row["dxmInternalOrderNo"] == 7821:
    #         continue  # Skip if dxmInternalOrderNo is 7821

    for i, row in enumerate(response):
        if len(str(row['invoiceM3'])) == 12:
            continue
        data = {
            "SlNo": str(i),
            "ReqNo": str(row["sl_NO"]),
            "DMXOrdNo": row["dxmordno"],
            "DMXInternal": row["dxmInternalOrderNo"],
            "DMXDate": row["tdate"],
            "UpdateStatus": row["dxmupdsts"],
            "ItemNo": row["m_ITEM_NO"],
            "CONo": row["m_ORNO"],
            "ItemError": row["m_ITEM_ERR"],
            "COError": row["m_ORNO_ERR"],
            "SalesPrice": row["salesPrice"],
            "FreightCharge": row["freightCharge"],
        }
        log(f"Processing CO: {data['CONo']}")
        log(f"DXMLOG: {data['DMXInternal']}")
        if row["content"] and row["styleDescription"]:
            data["Desc"] = row["content"] + " " + row["styleDescription"]
            log("Style Description is done!")
        print(data)
        if data["ItemNo"] and data["FreightCharge"] and data["SalesPrice"] and data["Desc"] and data["CONo"] and data["DMXOrdNo"] and not data_exists(data):
            try:
                log("HSN Update Started!")
                update_hsn(data["ItemNo"])
                log("HSN Update Over!")
            except:
                log("Error Updating HSN")
                give_error("HSN Error", f"Error While Updating HSN: \n{traceback.format_exc()}")
                continue
            try:
                log("Update CO Started")
                update_CO(data["CONo"], data["DMXOrdNo"], data["SalesPrice"])
                log("Update CO Ended")
            except:
                log("Error Updating CO")
                give_error("CO Error", f"Error Updating CO Details: \n{traceback.format_exc()}")
                continue
            try:
                log("Plan Creation Started")
                data["Plan"] = plan_creation(data["CONo"])
                log("Plan Creation Ended")
            except:
                log("Error Creating Plan")
                give_error("Plan Error", f"Error Generating Plan Number: \n{traceback.format_exc()}")
                continue
            try:
                log("Invoice Generation Started")
                upcharge = float(data["FreightCharge"]) * 100.0 / float(data["SalesPrice"])
                data["InvoiceDate"], data["CINumber"], data["InvoiceNo"], rosino = invoice(data["Plan"], upcharge, data["SalesPrice"], data["Desc"])
                log("Invoice Generation Ended")
            except:
                log(f"Error Generating Invoice, Plan: {data['Plan']}")
                give_error("Invoice Error", f"Error Generating Invoice, The plan number is: {data['Plan']}\nexception: {traceback.format_exc()}")
                print(traceback.format_exc())
                continue
            upload_data = {
                "invoiceM3": data["InvoiceNo"],
                "m3InvoiceDate": str(datetime.now().strftime("%Y-%m-%dT%H:%M:%S.000")),
                "slNo": int(data["ReqNo"]),
                "packType": "LSE-LOOSE PACK",
                "discriptionOfGoods": data["Desc"],
                "rosiNumber": rosino,
                "cifValue": 111.00,
                "serviceNumber": str(data["Plan"])
            }
            bod = json.dumps(upload_data)
            resp = requests.post(url="https://careers.shahi.co.in/shahiapiprods/apiTest/updateInvoiceNumber", data=bod, headers={"Content-Type": "application/json"})
            log("Updated To the API")
            print(resp.json())
            add_data(data)
            log("Updated To Google Sheet")


# invoice("5539205", 46.38, 69, "Something Something Description")
# update_CO("2000733141", "5351251673305-1-1", "69")

# plan_creation("2000733141")

# invoice_date, ci_num, invoice_no, rosino = invoice("", "0.46", "69", "")

def run_bot():
    try:
        bot()
    except Exception as e:
        log("Unforseen Error Occured")
        messagebox.showerror("Error", f"Application closed due to \n{traceback.format_exc()}")
    finally:
        run_bot()

#
# def update_label():
#     try:
#         status = status_queue.get(timeout=3)
#         label.config(text=status)
#     except queue.Empty:
#         print("Waiting")
#     finally:
#         root.after(1000, update_label)
#
#
# thread = Thread(target=run_bot, daemon=True)
# thread.start()
# root = tk.Tk()
# root.geometry("400x100")
# root.title("DXM Bot")
# label = tk.Label(root, text="DXM Bot Application Started")
# label.pack(pady=30)
# root.after(1000, update_label)
# root.mainloop()

# invoice('5560116', str(float(19) * 100.0 / float(60)), '60', '98% Cotton 2% Elastane Woven Mens Trousers')